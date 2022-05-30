#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
#  formatting_dynamics.py
#
#  Copyright 2022 a.sherchenkov <a.sherchenkov@WKS-LGT-DP17>
#
#  This program is free software; you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation; either version 2 of the License, or
#  (at your option) any later version.
#
#  This program is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#
#  You should have received a copy of the GNU General Public License
#  along with this program; if not, write to the Free Software
#  Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston,
#  MA 02110-1301, USA.
#
#
import os.path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, numbers
from openpyxl.utils import get_column_letter

def main(pathfile=os.path.join(os.getcwd(), "Python_modules", "UGS_module", "report", 'Динамика_прогнозных_значений.xlsx')):
    """ Форматирование динамики: удаление пустой строки,
    разряд 1000, округление до целого числа
    """

    wb = load_workbook(pathfile)
    ws = wb['ПХГ']

    ws.delete_rows(3)

    rigth_dawn_cell_name = ws.cell(ws.max_row, ws.max_column).coordinate

    for cellObj in ws['B2:{}'.format(rigth_dawn_cell_name)]:
        for cell in cellObj:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            # if type(cell.value).__name__  == 'float':
                # cell.value = int(cell.value)
            # cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

    # Выравнивание по ширине (20 символов)
    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].width = 20

    wb.save(pathfile)

    return 0

if __name__ == '__main__':
    import sys
    sys.exit(main())
