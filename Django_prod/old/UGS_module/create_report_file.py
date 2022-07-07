#!/usr/bin/env python

import shelve
from openpyxl import load_workbook
from openpyxl.styles.fills import PatternFill
from openpyxl.styles.colors import Color
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.utils import get_column_letter
import pandas as pd
import sys
import os


def paint_grey_massive(sheet, x0, y0, x1, y1):
    """ Раскраска серым цветом строк на листе!
    """

    _fill = PatternFill(patternType='solid', fgColor=Color(rgb='00C0C0C0'))
    for x in range(x0, x1 + 1):
        for y in range(y0, y1 + 1):
            sheet.cell(x, y).fill = _fill


def main():
    """ # Создание на основе итоговых отчетов агрегированной формы
    """

    try:
        with shelve.open('dbl', 'r') as db:
            df1 = db['report_by_destination']
            df2 = db['report_by_country']
    except FileNotFoundError:
        sys.exit("Файлы с данными не надены!")

    # Объединение 2 отчетов
    df = pd.concat([df2.append(pd.Series(dtype='float64'), ignore_index=True), df1], axis=0)
    temp_file = 'pre_report.xlsx'

    df.to_excel(temp_file, index=False)

    # Форматирование отчета
    wb = load_workbook(temp_file)
    ws = wb['Sheet1']

    rigth_dawn_cell_name = ws.cell(ws.max_row, ws.max_column).coordinate

    # Установка шрифта и размера
    for cellObj in ws['A1:{}'.format(rigth_dawn_cell_name)]:
        for cell in cellObj:
            cell.font = Font(name="Times New Roman", size=12)

    # Список номеров строк для серого цвета.
    # Серый цвет
    temp_grey = ['Показатель', 'Закачка газа, всего', 'Отбор газа, всего',
                'Велке Капушаны', 'Кондратки', 'Северный поток', 'Северный поток-2', 'Турецкий поток']
    grey = []
    for i in range(1, ws.max_row - 1):
        if ws.cell(i, 1).value in temp_grey:
            grey.append(i)

    # Заполнние фона жирным серым цветом
    # Номера колонок
    for y in [2, 6, 10, 14, 18, 19, 20]:
        paint_grey_massive(ws, 2, y, ws.max_row - 1, y)
    # Номера строк
    for x in grey[3:]:
        paint_grey_massive(ws, x, 1, x + 6, ws.max_column)

    # Жирный жирного шрифта яцеек, v2
    arr = range(1, ws.max_column + 1)
    row = grey
    col = [arr, arr, arr, [1], [1], [1], [1], [1]]
    for i in range(len(row)):
        for j in col[i]:
            ws.cell(row[i], j).font = Font(bold=True, name="Times New Roman", size=12)

    # Установка выравнивания ячеек, формат тысяч
    for cellObj in ws['B2:{}'.format(rigth_dawn_cell_name)]:
        for cell in cellObj:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

    # Выравнивание по ширине
    for idx, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(idx)].auto_size = True
    ws.column_dimensions[get_column_letter(1)].width = 50

    os.remove(temp_file)

    dir_name = 'report'
    if not os.path.exists(dir_name):
        os.mkdir(dir_name)

    try:
        wb.save(dir_name + '//Отчет.xlsx')
    except PermissionError as msg:
        sys.exit("Ошибка: Невозможно сохранить файл!")

    print("Агрегированный отчет сформирован!")

    return 0

if __name__  == "__main__":
    sys.exit(main())
